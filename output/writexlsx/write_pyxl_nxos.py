from openpyxl.styles import Font, Color, Alignment, borders, Border, Side, PatternFill, colors, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pprint import pprint
import math

font = Font(name='Calibri',
                            size=11,
                            bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FF000000')

white_font = Font(name='Calibri',
                            size=11,
                            bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FFFFFF')

fill = PatternFill(fill_type=None,
                    start_color='FFFFFFFF',
                    end_color='FF000000')
blue_fill = PatternFill(fill_type='solid',
                            start_color='0033CC',
                            end_color='0033CC')
light_green_fill = PatternFill(fill_type='solid',
                                start_color='C6EFCE',
                                end_color='C6EFCE')
light_blue_fill = PatternFill(fill_type='solid',
                                start_color='33CCFF',
                                end_color='33CCFF')
pink_fill = PatternFill(fill_type='solid',
                        start_color='FFC7CE',
                        end_color='FFC7CE')
border = Border(left=Side(border_style=borders.BORDER_THIN , color='000000'),
                right=Side(border_style=borders.BORDER_THIN, color='000000'),
                top=Side(border_style=borders.BORDER_THIN, color='000000'),
                bottom=Side(border_style=borders.BORDER_THIN, color='000000'))




class Headline:
    def __init__(self):
        # Key list is the json key, namelist is the spreadsheet headers
        self.keylist = []
        self.namelist = []


def getcolletter(index):
    letter = get_column_letter(index)
    return letter


def mkXLSX(filename, TAB='default'):
    # Tab will be first tab in WorkBook
    workbook = {}
    pprint('Creating Excel File: ' + filename)
    wb = Workbook()
    tab = wb.active
    tab.title = TAB
    return wb


def addSheet(WB, TAB):
    pprint('Creating Excel Tab: ' + TAB)
    newTab = WB.create_sheet(title=TAB)
    return newTab




#region Cell format functions
def mkPattern(fillType=None, startColor='FFFFFFFF', endColor='FF000000'):
    fill = PatternFill(fill_type=fillType,
                        start_color=startColor,
                        end_color=endColor)
    return fill

def mkFont(name='Calibri', size=11, bold=False, italic=False,
           vertAlign=None, underline='none', strike=False,
           color='FF000000'):
    font = Font(name=name,
                size=size,
                bold=bold,
                italic=italic,
                vertAlign=vertAlign,
                underline=underline,
                strike=strike,
                color=color)
    return font

def mkBorder(color='00000000'):
    border = Border(
        left=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color),
        right=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color),
        top=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color),
        bottom=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color))
    return border
#endregion
def setCellFormat(WS, ROW, COL, FILL=fill, BORDER=border, FONT=font):
    format_cell =  WS.cell(row=ROW, column=COL)
    format_cell.fill = FILL
    format_cell.border = BORDER
    format_cell.font = FONT


def write_hl(WS, NAMELIST, ROW, COL, FILL=light_blue_fill, BORDER=border, FONT=font): # Input Switch Object
    COL1 = COL
    ROW1 = ROW
    ROW1 = ROW1 + 1
    for h in NAMELIST:
        write_cell = WS.cell(row=ROW1, column=COL1)
        write_cell.value = h
        write_cell.fill = FILL
        write_cell.border = BORDER
        write_cell.font = FONT
        COL1 = COL1 + 1
    return ROW1 + 1

def write(WS, ROW, COL, DATA, BORDER=border, FONT=font): # Input Switch Object
    COL1 = COL
    ROW1 = ROW
    write_cell = WS.cell(row=ROW1, column=COL1)
    write_cell.value = DATA
    write_cell.border = BORDER
    write_cell.font = FONT


def write_obj(WS, OBJ_List, KEYLIST, ROW, COL): # Input Object to write
    COL1 = COL
    ROW1 = ROW
    for a in OBJ_List:
        ROW1 = write_line(WS, a, KEYLIST, ROW1, COL1)
    return ROW1

def write_fill(WS, ROW, COL, DATA, FILL=fill, BORDER=border, FONT=font): # Input Switch Object
    COL1 = COL
    ROW1 = ROW
    write_cell = WS.cell(row=ROW1, column=COL1)
    write_cell.value = DATA
    write_cell.border = BORDER
    write_cell.font = FONT
    write_cell.fill = FILL

def write_line(WS, OBJ, KEYLIST, ROW, COL, FILL=fill):
    COL1 = COL
    ROW1 = ROW
    for i in KEYLIST:
        try:
            b = getattr(OBJ, i)
        except:
            b = ''
        write_fill(WS, ROW1, COL1, b, FILL)
        COL1 = COL1 + 1
    return ROW1 + 1

def write_portmap(WS, OBJ, KEYLIST, ROW, COL):
    COL1 = COL
    ROW1 = ROW
    color = fill
    pprint('write portmap')
    for i in KEYLIST:
        try:
            b = getattr(OBJ, i)
        except:
            b = ''
        if i == 'status':
            if b == 'connected' or b == 'up/up':
                color = light_green_fill
            else:
                color = pink_fill
        write_fill(WS, ROW1, COL1, b, color)
        color = fill
        COL1 = COL1 + 1
        pprint(b)
    return ROW1 + 1


def write_merge(WS, startROW, startCOL, endROW, endCOL, DATA):
    write(WS, startROW, startCOL, DATA)
    WS.merge_cells(start_row=startROW, start_column=startCOL, end_row=endROW, end_column=endCOL - 1)
    return endROW + 1


def write_hl_category(WS, ROW, COL, category, hl, endCol=1):
    hl_len = len(hl)
    if hl_len > endCol:
        endCol = hl_len
    ROW = ROW + 2
    endCol = COL + endCol - 1
    ROW = write_merge(WS, ROW, COL, ROW, endCol, category)
    ROW = ROW - 1
    ROW = write_hl(WS, hl, ROW, COL)
    return ROW


def write_family_hl(WS, parent, family, keys, ROW, COL, D_FORMAT):
    length = len(family)
    for a in parent:
            aa = family[0]
            ### ROW = write_data.write_line(wsSystem, d, bgp, ROW, COL, D_FORMAT)
            for b in getattr(a, aa):
                if length > 1:
                    bb = family[1]
                    if hasattr(c, cc):
                        for c in getattr(b, bb):
                            if length > 2:
                                cc = family[2]
                                if hasattr(c, cc):
                                    for d in getattr(c, cc):
                                        if length > 3:
                                            dd = family[3]
                                            if hasattr(d, dd):
                                                for e in getattr(d, dd):
                                                    if length > 4:
                                                        ee = family[4]
                                                        if hasattr(d, dd):
                                                            for e in getattr(d, dd):
                                                                ROW = write_line(WS, e, keys, ROW, COL,
                                                                                            D_FORMAT)

                                                    else:
                                                        ROW = write_line(WS, e, keys, ROW, COL, D_FORMAT)
                                        else:
                                            ROW = write_line(WS, d, keys, ROW, COL, D_FORMAT)
                            else:
                                ROW = write_line(WS, c, keys, ROW, COL, D_FORMAT)
                else:
                    ROW = write_line(WS, b, keys, ROW, COL, D_FORMAT)

    return ROW



def setalignment(WS, ROW, COL, horizontal='center', vertical='center'):
    write_cell = WS.cell(row=ROW, column=COL)
    write_cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

def setfontcolor(WS, ROW, COL, COLOR):
    write_cell = WS.cell(row=ROW, column=COL)
    write_cell.font = Font(color=COLOR)

def setfill(WS, ROW, COL, COLOR):
    write_cell = WS.cell(row=ROW, column=COL)
    write_cell.fill = PatternFill(fill_type='solid',
                                  start_color=COLOR,
                                  end_color=COLOR)

def setformatcolumn(WS, ROW, COL, FORMAT):
    WS.write(ROW, COL, VALUE, FORMAT)


