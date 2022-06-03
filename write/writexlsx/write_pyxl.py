from openpyxl.styles import Font, Color, Alignment, borders, Border, Side, PatternFill, colors
from openpyxl import Workbook
from pprint import pprint
import math

font = Font(name='Calibri',
                            size=11,
                            bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FF000000')

white_font = Font(name='Calibri',
                            size=11,
                            bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FFFFFF')

fill = PatternFill(fill_type=None,
                    start_color='FFFFFFFF',
                    end_color='FF000000')
green_fill = PatternFill(fill_type='solid',
                            start_color='00B050',
                            end_color='00B050')
light_green_fill = PatternFill(fill_type='solid',
                                start_color='C6EFCE',
                                end_color='C6EFCE')
pink_fill = PatternFill(fill_type='solid',
                        start_color='FFC7CE',
                        end_color='FFC7CE')
border = Border(left=Side(border_style=borders.BORDER_THIN , color='000000'),
                right=Side(border_style=borders.BORDER_THIN, color='000000'),
                top=Side(border_style=borders.BORDER_THIN, color='000000'),
                bottom=Side(border_style=borders.BORDER_THIN, color='000000'))




class Headline:
    def __init__(self):
        self.keylist = ['id', 'speed', 'adminSt', 'operSt', 'PG', 'descr', 'selDescr', 'layer', 'autoNeg', 'mtu', 'portT', 'guiCiscoPID', 'guiSN', 'guiName',
                        'devId', 'portId', 'platId', 'ver', 'sysName', 'portDesc']
        self.namelist = ['Interface', 'Speed','Admin State', 'Operational State', 'Port Group', 'Interface Desc', 'Int Selector Desc', 'Layer', 'Auto Neg', 'MTU',
                         'Port Type', 'Transceiver', 'Transceiver SN', 'Vender', 'CDP Neighbor', 'CDP Port', 'Platform',
                         'Software Version', 'LLDP Neighbor', 'LLDP Port']

def mkXLSX(fileName, TAB='default'):
    # Tab will be first tab in WorkBook
    workbook = {}
    pprint('Creating Excel File: ' + fileName)
    wb = Workbook()
    tab = wb.active
    tab.title = TAB
    return wb
def addSheet(WB, TAB):
    pprint('Creating Excel Tab: ' + TAB)
    newTab = WB.create_sheet(title=TAB)
    return newTab




#region Cell format functions
def mkPattern(fillType=None, startColor='FFFFFFFF', endColor='FF000000'):
    fill = PatternFill(fill_type=fillType,
                        start_color=startColor,
                        end_color=endColor)
    return fill

def mkFont(name='Calibri', size=11, bold=False, italic=False,
           vertAlign=None, underline='none', strike=False,
           color='FF000000'):
    font = Font(name=name,
                size=size,
                bold=bold,
                italic=italic,
                vertAlign=vertAlign,
                underline=underline,
                strike=strike,
                color=color)
    return font

def mkBorder(color='00000000'):
    border = Border(
        left=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color),
        right=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color),
        top=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color),
        bottom=Side(border_style=openpyxl.styles.borders.BORDER_THIN, color=color))
    return border
#endregion
def setCellFormat(WS, ROW, COL, FILL=fill, BORDER=border, FONT=font):
    format_cell =  WS.cell(row=ROW, column=COL)
    format_cell.fill = FILL
    format_cell.border = BORDER
    format_cell.font = FONT


def write_hl(WS, NAMELIST, ROW, COL, FILL=green_fill, BORDER=border, FONT=font): # Input Switch Object
    COL1 = COL
    ROW1 = ROW
    ROW1 = ROW1 + 1
    for h in NAMELIST:
        write_cell = WS.cell(row=ROW1, column=COL1)
        write_cell.value = h
        write_cell.fill = FILL
        write_cell.border = BORDER
        write_cell.font = FONT
        COL1 = COL1 + 1
    return ROW1 + 1

def write(WS, ROW, COL, DATA, BORDER=border, FONT=font): # Input Switch Object
    COL1 = COL
    ROW1 = ROW
    write_cell = WS.cell(row=ROW1, column=COL1)
    write_cell.value = DATA
    write_cell.border = BORDER
    write_cell.font = FONT


def write_obj(WS, OBJ, keylist, ROW, COL): # Input Interface Object for OBJ and Switch Object for HL
    COL1 = COL
    ROW1 = ROW
    for a in keylist:
        try:
            b = getattr(OBJ, a)
            WS.write(ROW1, COL1, b)
        except:
            print(a + ' Does not exist in ' + OBJ.name)
        finally:
            COL1 = COL1 + 1
    return ROW1 + 1

def write_fill(WS, ROW, COL, DATA, FILL=fill, BORDER=border, FONT=font): # Input Switch Object
    COL1 = COL
    ROW1 = ROW
    write_cell = WS.cell(row=ROW1, column=COL1)
    write_cell.value = DATA
    write_cell.border = BORDER
    write_cell.font = FONT
    write_cell.fill = FILL

def write_line(WS, OBJ, KEYLIST, ROW, COL, FILL=fill):
    COL1 = COL
    ROW1 = ROW
    for i in KEYLIST:
        try:
            b = getattr(OBJ, i)
        except:
            b = ''
        write_fill(WS, ROW1, COL1, b, FILL)
        COL1 = COL1 + 1
    return ROW1 + 1

def write_portmap(WS, OBJ, KEYLIST, ROW, COL):
    COL1 = COL
    ROW1 = ROW
    color = fill
    pprint('write portmap')
    for i in KEYLIST:
        try:
            b = getattr(OBJ, i)
        except:
            b = ''
        if i == 'status':
            if b == 'connected' or b == 'up/up':
                color = light_green_fill
            else:
                color = pink_fill
        write_fill(WS, ROW1, COL1, b, color)
        color = fill
        COL1 = COL1 + 1
        pprint(b)
    return ROW1 + 1


def write_merge(WS, startROW, startCOL, endROW, endCOL, DATA, FILL=green_fill, BORDER=border, FONT=white_font):
    write_fill(WS, startROW, startCOL, DATA, FILL, BORDER, FONT)
    WS.merge_cells(start_row=startROW, start_column=startCOL, end_row=endROW, end_column=endCOL)
    return endROW + 1


def write_hl_category(WS, ROW, COL, category, hl, endCol=1):
    hl_len = len(hl)
    if hl_len > endCol:
        endCol = hl_len
    ROW = ROW + 2
    endCol = COL + endCol - 1
    ROW = write_merge(WS, ROW, COL, ROW, endCol, category)
    ROW = ROW - 1
    ROW = write_hl(WS, hl, ROW, COL)
    return ROW


def write_family_hl(WS, parent, family, keys, ROW, COL, D_FORMAT):
    length = len(family)
    for a in parent:
            aa = family[0]
            ### ROW = write_data.write_line(wsSystem, d, bgp, ROW, COL, D_FORMAT)
            for b in getattr(a, aa):
                if length > 1:
                    bb = family[1]
                    if hasattr(c, cc):
                        for c in getattr(b, bb):
                            if length > 2:
                                cc = family[2]
                                if hasattr(c, cc):
                                    for d in getattr(c, cc):
                                        if length > 3:
                                            dd = family[3]
                                            if hasattr(d, dd):
                                                for e in getattr(d, dd):
                                                    if length > 4:
                                                        ee = family[4]
                                                        if hasattr(d, dd):
                                                            for e in getattr(d, dd):
                                                                ROW = write_line(WS, e, keys, ROW, COL,
                                                                                            D_FORMAT)

                                                    else:
                                                        ROW = write_line(WS, e, keys, ROW, COL, D_FORMAT)
                                        else:
                                            ROW = write_line(WS, d, keys, ROW, COL, D_FORMAT)
                            else:
                                ROW = write_line(WS, c, keys, ROW, COL, D_FORMAT)
                else:
                    ROW = write_line(WS, b, keys, ROW, COL, D_FORMAT)

    return ROW

def setformatcolumn(WS, ROW, COL, FORMAT):
    WS.write(ROW, COL, VALUE, FORMAT)